from pathlib import Path
from openpyxl import load_workbook
from ..members.models import Member
from .column_mapper import map_headers


def import_xlsx(path: Path) -> list[Member]:
    wb=load_workbook(path,read_only=True,data_only=True); ws=wb.active
    it=ws.iter_rows(values_only=True); headers=[str(x or '') for x in next(it)]; mapping=map_headers(headers); out=[]
    for row in it:
        data={field:(row[i] if i < len(row) else None) for i,field in mapping.items()}
        raw_id=data.get('telegram_user_id')
        try: uid=int(raw_id) if raw_id not in (None,'') else None
        except (ValueError,TypeError): uid=None
        out.append(Member(uid,str(data.get('username') or '') or None,str(data.get('first_name') or '') or None,str(data.get('last_name') or '') or None,str(data.get('phone') or '') or None,bool(data.get('bot',False)),bool(data.get('deleted',False)),str(data.get('activity_status') or '') or None,str(data.get('last_seen') or '') or None))
    wb.close(); return out
