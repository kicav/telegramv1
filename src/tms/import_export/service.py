from __future__ import annotations
import csv
from pathlib import Path
from ..core.enums import JobState,JobType
from ..core.events import DomainEvent
from ..datasets.models import Dataset
from ..jobs.models import Job
from ..members.models import Member

class ImportExportService:
    def __init__(self,datasets,members,jobs,events)->None:self.datasets=datasets;self.members=members;self.jobs=jobs;self.events=events
    @staticmethod
    def _int(v):
        try:return int(str(v).strip()) if v not in (None,'') else None
        except (ValueError,TypeError):return None
    def import_dataset(self,path:Path,name:str,account_id:int|None=None):
        path=Path(path);dataset_id=self.datasets.create(Dataset(None,name or path.stem,'FILE',str(path)));job_id=self.jobs.create(Job(None,JobType.IMPORT,JobState.RUNNING,account_id=account_id,source_dataset_id=dataset_id));rows=[]
        if path.suffix.lower()=='.csv':
            with path.open('r',encoding='utf-8-sig',newline='') as f:data=list(csv.DictReader(f))
        elif path.suffix.lower()=='.xlsx':
            from openpyxl import load_workbook
            wb=load_workbook(path,read_only=True,data_only=True);ws=wb.active;it=ws.iter_rows(values_only=True);headers=[str(x or '').strip() for x in next(it,())];data=[dict(zip(headers,r)) for r in it];wb.close()
        else:raise ValueError('Only CSV/XLSX are supported')
        invalid=0
        for r in data:
            uid=self._int(r.get('user_id') or r.get('telegram_user_id'));username=(str(r.get('username') or '').strip().lstrip('@') or None);access=self._int(r.get('access_hash'))
            if uid is None and not username:invalid+=1;continue
            rows.append(Member(uid,username,str(r.get('first_name') or '') or None,str(r.get('last_name') or '') or None,str(r.get('phone') or '') or None,access_hash=access))
        summary=self.members.submit_ingest_batch(dataset_id,rows,account_id=account_id,source_label=path.name).result(timeout=60)
        state=JobState.COMPLETED_WITH_ERRORS if invalid+summary.invalid else JobState.COMPLETED
        self.jobs.submit_set_state(job_id,state).result(timeout=10);self.events.publish(DomainEvent('ImportCompleted',{'dataset_id':dataset_id,'job_id':job_id}));return dataset_id,job_id
    def export_dataset(self,dataset_id:int,path:Path,account_id:int|None=None):
        job_id=self.jobs.create(Job(None,JobType.EXPORT,JobState.RUNNING,account_id=account_id,source_dataset_id=dataset_id));path=Path(path);path.parent.mkdir(parents=True,exist_ok=True);headers=['user_id','access_hash','username','first_name','last_name','phone','bot','deleted','activity_status','last_seen']
        def row_values(m):
            access=None
            if account_id is not None and m.telegram_user_id is not None:
                with self.members.db.reader() as conn:r=conn.execute('SELECT access_hash FROM peer_cache WHERE account_id=? AND peer_id=?',(account_id,m.telegram_user_id)).fetchone();access=r[0] if r else None
            return [m.telegram_user_id,access,m.username,m.first_name,m.last_name,m.phone,int(m.bot),int(m.deleted),m.activity_status,m.last_seen]
        if path.suffix.lower()=='.xlsx':
            from openpyxl import Workbook
            wb=Workbook(write_only=True);ws=wb.create_sheet('Members');ws.append(headers)
            for chunk in self.members.iter_dataset_rows(dataset_id):
                for m in chunk:ws.append(row_values(m))
            wb.save(path)
        else:
            with path.open('w',encoding='utf-8-sig',newline='') as f:
                w=csv.writer(f);w.writerow(headers)
                for chunk in self.members.iter_dataset_rows(dataset_id):
                    for m in chunk:w.writerow(row_values(m))
        self.jobs.submit_set_state(job_id,JobState.COMPLETED).result(timeout=10);return job_id
    def export_job_results(self,job_id:int,path:Path):
        path=Path(path);path.parent.mkdir(parents=True,exist_ok=True)
        with self.members.db.reader() as conn:rows=conn.execute('''SELECT mi.ordinal,m.telegram_user_id,m.username,mi.state,mi.attempt_count,mi.last_error_code,mi.last_error_text FROM migration_items mi JOIN members m ON m.id=mi.member_id WHERE mi.job_id=? ORDER BY mi.ordinal''',(job_id,)).fetchall()
        with path.open('w',encoding='utf-8-sig',newline='') as f:w=csv.writer(f);w.writerow(['ordinal','telegram_user_id','username','state','attempt_count','error_code','error_text']);w.writerows([tuple(r) for r in rows])
    def export_job_log(self,job_id:int,path:Path):
        with self.members.db.reader() as conn:rows=conn.execute('SELECT timestamp,level,event_code,member_id,message FROM job_events WHERE job_id=? ORDER BY id',(job_id,)).fetchall()
        with Path(path).open('w',encoding='utf-8-sig',newline='') as f:w=csv.writer(f);w.writerow(['timestamp','level','event_code','member_id','message']);w.writerows([tuple(r) for r in rows])
